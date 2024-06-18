from openpyxl import Workbook, load_workbook
import json

"""For now file already converted, but may be useful in the future"""


def fill_countries(filename):
    """Fill column of countries"""
    wb = load_workbook(filename)
    sheet = wb['Table 1']
    ws = wb.active

    # fill columns
    columns = ['D', 'C', 'B', 'A']
    for col in range(3):
        i = 1
        cell = ws[f'{columns[col + 1]}{i}'].value
        while ws[f'{columns[col]}{i}'].value:
            if ws[f'{columns[col + 1]}{i}'].value:
                cell = ws[f'{columns[col + 1]}{i}'].value
            else:
                ws[f'{columns[col + 1]}{i}'].value = cell
            i += 1

    wb.save(filename)


# fill_countries('airports.xlsx')


def convert_to_json(filename):
    wb = load_workbook(filename)
    sheet = wb['Table 1']
    ws = wb.active

    data = []
    for row in ws.iter_rows():
        data.append({
            'country': row[0].value,
            'code': '',
            'city': row[1].value,
            'name': row[2].value,
            'service': row[3].value,
            'arrival': {
                'price_one_passanger': row[4].value,
                'price_other_passangers': row[5].value
            },
            'departure': {
                'price_one_passanger': row[6].value,
                'price_other_passangers': row[7].value
            },
            'transit': {
                'price_one_passanger': row[8].value,
                'price_other_passangers': row[9].value
            },
            'additional_info': row[10].value
        })

    with open('airports_json.json', 'w', encoding='utf-8') as f:

        json.dump(data, f, ensure_ascii=False, indent=4)


# convert_to_json('airports.xlsx')

# s = 'Andorra-La Seu dвЂ™Urgell Airport (LEU)'
# s1 = s.encode('cp1251').decode('utf-8')
# print(s1)


def decode_csv(filename):
    """Fix csv encoding"""
    with (open(filename, 'r', encoding='utf-8') as f):
        data = f.read().encode('cp1251', errors='ignore')
        # print(data)
        data = data.decode('utf-8', errors='ignore')
    with open('new.csv', 'w', encoding='utf-8') as f:
        f.write(data)


# decode_csv('airport_airport.csv')
# decode_csv('main_app_service.csv')
